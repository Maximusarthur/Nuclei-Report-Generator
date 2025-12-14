#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IP漏洞报告生成器（完整版）
包含GUI界面和核心处理逻辑
"""

import re
import pandas as pd
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
from datetime import datetime
import subprocess


class IPReportGenerator:
    """IP漏洞报告生成器（完整版）"""

    def __init__(self, root, main_app):
        self.root = root
        self.main_app = main_app
        self.root.title("IP漏洞报告生成器")
        self.root.geometry("900x700")

        # 设置窗口图标
        try:
            self.root.iconbitmap(default="nuclei.ico")
        except:
            pass

        # 初始化变量
        self.target_files = []
        self.scan_files = []

        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        # 返回首页按钮
        back_frame = tk.Frame(self.root)
        back_frame.pack(fill="x", padx=10, pady=5)

        tk.Button(back_frame, text="← 返回首页",
                  font=("微软雅黑", 10), bg="#95a5a6", fg="white",
                  command=self.return_to_main).pack(anchor="w")

        # 标题
        title_label = tk.Label(self.root, text="IP漏洞报告生成器",
                               font=("微软雅黑", 18, "bold"), fg="#2c3e50")
        title_label.pack(pady=10)

        # 说明标签
        desc_text = """• 目标文件和扫描结果文件按顺序一一对应
• 例如：安全设备.txt -> 安全设备.xlsx
• 文件数量必须相同且顺序匹配"""
        desc_label = tk.Label(self.root, text=desc_text, font=("微软雅黑", 10),
                              justify="left", fg="#34495e", bg="#ecf0f1", wraplength=850)
        desc_label.pack(pady=5, padx=20, fill="x")

        # 创建主框架 - 使用PanedWindow实现可调整分割
        main_paned = tk.PanedWindow(self.root, orient=tk.VERTICAL, sashrelief=tk.RAISED)
        main_paned.pack(fill="both", expand=True, padx=10, pady=5)

        # 上半部分：文件选择和输出设置
        top_frame = tk.Frame(main_paned)
        main_paned.add(top_frame, minsize=400)

        # 使用Notebook标签页优化空间利用
        style = ttk.Style()
        style.configure('Custom.TNotebook.Tab',
                        font=('微软雅黑', 12, 'bold'),
                        padding=[15, 8],
                        background='#ecf0f1')

        style.map('Custom.TNotebook.Tab',
                  background=[('selected', '#3498db')],
                  foreground=[('selected', 'white')])

        notebook = ttk.Notebook(top_frame, style='Custom.TNotebook')
        notebook.pack(fill="both", expand=True, padx=5, pady=5)

        # 标签页1：文件选择 - 使用彩色边框
        self.file_content = tk.Frame(notebook, bg='white', highlightbackground='#3498db', highlightthickness=2)
        notebook.add(self.file_content, text="📁 文件选择")

        # 标签页2：输出和控制 - 使用彩色边框
        self.control_content = tk.Frame(notebook, bg='white', highlightbackground='#2ecc71', highlightthickness=2)
        notebook.add(self.control_content, text="⚙️ 输出控制")

        # 创建文件选择内容
        self.create_file_content()

        # 创建输出控制内容
        self.create_control_content()

        # 下半部分：进度条和日志
        bottom_frame = tk.Frame(main_paned)
        main_paned.add(bottom_frame, minsize=200)

        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(bottom_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", padx=10, pady=(10, 5))

        # 日志文本框
        log_frame = tk.LabelFrame(bottom_frame, text="处理日志", font=("微软雅黑", 11, "bold"))
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.log_text = scrolledtext.ScrolledText(log_frame, height=6, font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, padx=10, pady=10)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = tk.Label(self.root, textvariable=self.status_var,
                              font=("微软雅黑", 9), bg="#34495e", fg="white",
                              anchor="w", padx=10)
        status_bar.pack(side="bottom", fill="x")

        # 设置窗口最小尺寸
        self.root.minsize(800, 600)

        # 添加日志
        self.log("IP漏洞报告生成器启动成功")
        self.log(f"当前工作目录: {os.getcwd()}")

    def create_file_content(self):
        """创建文件选择内容"""
        # 左侧目标文件，右侧扫描文件
        file_select_frame = tk.Frame(self.file_content)
        file_select_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 目标文件选择
        target_frame = tk.LabelFrame(file_select_frame, text="目标列表文件(.txt)",
                                     font=("微软雅黑", 10, "bold"))
        target_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # 目标文件列表框
        self.target_listbox = tk.Listbox(target_frame, height=8, font=("Consolas", 9))
        self.target_listbox.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        target_scroll = tk.Scrollbar(target_frame)
        target_scroll.pack(side="right", fill="y")
        self.target_listbox.config(yscrollcommand=target_scroll.set)
        target_scroll.config(command=self.target_listbox.yview)

        # 目标文件按钮
        target_btn_frame = tk.Frame(target_frame)
        target_btn_frame.pack(fill="x", padx=5, pady=5)

        tk.Button(target_btn_frame, text="添加", command=self.add_target_file,
                  font=("微软雅黑", 9), bg="#3498db", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(target_btn_frame, text="移除", command=self.remove_selected_target,
                  font=("微软雅黑", 9), bg="#e74c3c", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(target_btn_frame, text="清空", command=self.clear_target_list,
                  font=("微软雅黑", 9), bg="#95a5a6", fg="white", width=8).pack(side="left", padx=2)

        # 扫描文件选择
        scan_frame = tk.LabelFrame(file_select_frame, text="扫描结果文件(.txt)",
                                   font=("微软雅黑", 10, "bold"))
        scan_frame.pack(side="right", fill="both", expand=True, padx=(5, 0))

        # 扫描文件列表框
        self.scan_listbox = tk.Listbox(scan_frame, height=8, font=("Consolas", 9))
        self.scan_listbox.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        scan_scroll = tk.Scrollbar(scan_frame)
        scan_scroll.pack(side="right", fill="y")
        self.scan_listbox.config(yscrollcommand=scan_scroll.set)
        scan_scroll.config(command=self.scan_listbox.yview)

        # 扫描文件按钮
        scan_btn_frame = tk.Frame(scan_frame)
        scan_btn_frame.pack(fill="x", padx=5, pady=5)

        tk.Button(scan_btn_frame, text="添加", command=self.add_scan_file,
                  font=("微软雅黑", 9), bg="#3498db", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(scan_btn_frame, text="移除", command=self.remove_selected_scan,
                  font=("微软雅黑", 9), bg="#e74c3c", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(scan_btn_frame, text="清空", command=self.clear_scan_list,
                  font=("微软雅黑", 9), bg="#95a5a6", fg="white", width=8).pack(side="left", padx=2)

        # 文件对应提示
        self.pair_label = tk.Label(self.file_content, text="📝 请添加文件",
                                   font=("微软雅黑", 9), fg="#95a5a6")
        self.pair_label.pack(anchor="w", pady=(10, 5), padx=15)

        # 绑定列表变化事件
        self.target_listbox.bind('<<ListboxSelect>>', self.update_pair_status)
        self.scan_listbox.bind('<<ListboxSelect>>', self.update_pair_status)

    def create_control_content(self):
        """创建输出控制内容"""
        # 输出目录选择
        tk.Label(self.control_content, text="输出目录:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(10, 5), padx=15)

        output_frame = tk.Frame(self.control_content)
        output_frame.pack(fill="x", padx=15, pady=5)

        self.output_dir_var = tk.StringVar(value=os.path.join(os.getcwd(), "ip_reports"))
        output_entry = tk.Entry(output_frame, textvariable=self.output_dir_var,
                                font=("微软雅黑", 9))
        output_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))

        tk.Button(output_frame, text="浏览", command=self.select_output_dir,
                  font=("微软雅黑", 9), bg="#2ecc71", fg="white", width=8).pack(side="right")

        # 检测时间输入
        tk.Label(self.control_content, text="检测时间:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(20, 5), padx=15)

        # 使用当前日期作为默认值
        default_date = datetime.now().strftime("%Y年%m月%d日")
        self.scan_date_var = tk.StringVar(value=default_date)

        date_frame = tk.Frame(self.control_content)
        date_frame.pack(fill="x", padx=15, pady=5)

        date_entry = tk.Entry(date_frame, textvariable=self.scan_date_var,
                              font=("微软雅黑", 9))
        date_entry.pack(side="left", fill="x", expand=True)

        tk.Label(date_frame, text="格式：YYYY年MM月DD日",
                 font=("微软雅黑", 8), fg="#7f8c8d").pack(side="right", padx=(5, 0))

        # 版本号输入
        tk.Label(self.control_content, text="Nuclei版本号:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(15, 5), padx=15)

        self.tool_version_var = tk.StringVar(value="V3.5.1,nuclei-templates v10.3.2")

        version_frame = tk.Frame(self.control_content)
        version_frame.pack(fill="x", padx=15, pady=5)

        version_entry = tk.Entry(version_frame, textvariable=self.tool_version_var,
                                 font=("微软雅黑", 9))
        version_entry.pack(side="left", fill="x", expand=True)

        tk.Label(version_frame, text="示例：V3.5.1,nuclei-templates v10.3.2",
                 font=("微软雅黑", 8), fg="#7f8c8d").pack(side="right", padx=(5, 0))

        # 报告类型选择
        tk.Label(self.control_content, text="报告格式:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(20, 5), padx=15)

        self.report_type_var = tk.StringVar(value="excel")

        report_frame = tk.Frame(self.control_content)
        report_frame.pack(fill="x", padx=15, pady=5)

        tk.Radiobutton(report_frame, text="Word文档(.docx)",
                       variable=self.report_type_var, value="word",
                       font=("微软雅黑", 9)).pack(side="left", padx=(0, 15))

        tk.Radiobutton(report_frame, text="Excel表格(.xlsx)",
                       variable=self.report_type_var, value="excel",
                       font=("微软雅黑", 9)).pack(side="left")

        # 处理选项
        tk.Label(self.control_content, text="处理选项:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(20, 5), padx=15)

        self.auto_open_var = tk.BooleanVar(value=True)
        tk.Checkbutton(self.control_content, text="处理后自动打开报告文件",
                       variable=self.auto_open_var, font=("微软雅黑", 9)).pack(anchor="w", padx=15)

        # 处理按钮
        btn_frame = tk.Frame(self.control_content)
        btn_frame.pack(fill="x", pady=30, padx=15)

        self.process_btn = tk.Button(btn_frame, text="开始生成报告 (一一对应)",
                                     command=self.start_processing,
                                     font=("微软雅黑", 12, "bold"), bg="#27ae60", fg="white",
                                     padx=20, pady=10, state=tk.NORMAL)
        self.process_btn.pack(fill="x", pady=5)

    def return_to_main(self):
        """返回主界面"""
        self.root.destroy()
        self.main_app.return_to_home()

    def add_target_file(self):
        """添加目标文件"""
        files = filedialog.askopenfilenames(
            title="选择目标列表文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if files:
            for file in files:
                if file not in self.target_files:
                    self.target_files.append(file)
                    self.target_listbox.insert(tk.END, os.path.basename(file))
                    self.log(f"添加目标文件: {os.path.basename(file)}")
            self.update_pair_status()

    def remove_selected_target(self):
        """移除选中的目标文件"""
        selection = self.target_listbox.curselection()
        if selection:
            index = selection[0]
            removed_file = self.target_files.pop(index)
            self.target_listbox.delete(index)
            self.log(f"移除目标文件: {os.path.basename(removed_file)}")
            self.update_pair_status()

    def clear_target_list(self):
        """清空目标文件列表"""
        self.target_files.clear()
        self.target_listbox.delete(0, tk.END)
        self.log("清空目标文件列表")
        self.update_pair_status()

    def add_scan_file(self):
        """添加扫描文件"""
        files = filedialog.askopenfilenames(
            title="选择扫描结果文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if files:
            for file in files:
                if file not in self.scan_files:
                    self.scan_files.append(file)
                    self.scan_listbox.insert(tk.END, os.path.basename(file))
                    self.log(f"添加扫描文件: {os.path.basename(file)}")
            self.update_pair_status()

    def remove_selected_scan(self):
        """移除选中的扫描文件"""
        selection = self.scan_listbox.curselection()
        if selection:
            index = selection[0]
            removed_file = self.scan_files.pop(index)
            self.scan_listbox.delete(index)
            self.log(f"移除扫描文件: {os.path.basename(removed_file)}")
            self.update_pair_status()

    def clear_scan_list(self):
        """清空扫描文件列表"""
        self.scan_files.clear()
        self.scan_listbox.delete(0, tk.END)
        self.log("清空扫描文件列表")
        self.update_pair_status()

    def update_pair_status(self, event=None):
        """更新文件对应关系状态"""
        target_count = len(self.target_files)
        scan_count = len(self.scan_files)

        if target_count == scan_count:
            if target_count == 0:
                self.pair_label.config(text="📝 请添加文件", fg="#95a5a6")
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.pair_label.config(text=f"✅ 文件对应关系正常 ({target_count} 对文件)", fg="#27ae60")
                self.process_btn.config(state=tk.NORMAL)
        else:
            self.pair_label.config(
                text=f"⚠️ 文件数量不匹配: 目标文件 {target_count} 个，扫描文件 {scan_count} 个",
                fg="#e74c3c"
            )
            self.process_btn.config(state=tk.DISABLED)

    def select_output_dir(self):
        """选择输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir_var.set(directory)
            self.log(f"设置输出目录: {directory}")

    def log(self, message):
        """添加日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_status(self, message):
        """更新状态栏"""
        self.status_var.set(message)
        self.root.update_idletasks()

    def update_progress(self, value):
        """更新进度条"""
        self.progress_var.set(value)
        self.root.update_idletasks()

    def start_processing(self):
        """开始批量处理"""
        # 检查文件数量是否匹配
        if len(self.target_files) != len(self.scan_files):
            messagebox.showerror(
                "错误",
                f"文件数量不匹配！\n\n"
                f"目标文件数: {len(self.target_files)}\n"
                f"扫描文件数: {len(self.scan_files)}\n\n"
                f"请确保两个列表的数量相同且顺序对应。"
            )
            return

        if not self.target_files or not self.scan_files:
            messagebox.showwarning("警告", "请至少添加一对文件")
            return

        # 检查输出目录
        output_dir = self.output_dir_var.get()
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except Exception as e:
                messagebox.showerror("错误", f"无法创建输出目录: {e}")
                return

        # 在后台线程中处理
        processing_thread = threading.Thread(target=self.process_files_one_to_one)
        processing_thread.daemon = True
        processing_thread.start()

    def process_files_one_to_one(self):
        """一一对应处理文件"""
        total_pairs = len(self.target_files)

        self.update_status("开始处理文件...")
        self.update_progress(0)
        self.log(f"开始一一对应处理 {total_pairs} 对文件")

        success_count = 0
        failed_count = 0

        for i in range(total_pairs):
            try:
                target_file = self.target_files[i]
                scan_file = self.scan_files[i]

                # 计算进度
                progress = ((i + 1) / total_pairs) * 100
                self.update_progress(progress)

                # 获取文件名
                target_basename = os.path.splitext(os.path.basename(target_file))[0]
                scan_basename = os.path.splitext(os.path.basename(scan_file))[0]

                # 生成输出文件名
                report_type = self.report_type_var.get()
                if report_type == "word":
                    output_filename = f"{target_basename}-安全检测报告.docx"
                else:
                    output_filename = f"{target_basename}-report.xlsx"

                output_path = os.path.join(self.output_dir_var.get(), output_filename)

                # 记录处理开始
                pair_info = f"配对 {i + 1}/{total_pairs}: {os.path.basename(target_file)} -> {os.path.basename(scan_file)}"
                self.log(f"开始处理: {pair_info}")
                self.update_status(f"处理中: {output_filename}")

                # 调用处理函数
                if report_type == "word":
                    success = self.generate_word_report(target_file, scan_file, output_path)
                else:
                    success = self.generate_excel_report(target_file, scan_file, output_path)

                if success:
                    success_count += 1
                    self.log(f"✓ 处理完成: {output_filename}")

                    # 自动打开文件
                    if self.auto_open_var.get():
                        self.open_file(output_path)
                else:
                    failed_count += 1
                    self.log(f"✗ 处理失败: {os.path.basename(scan_file)}")

            except Exception as e:
                failed_count += 1
                self.log(f"✗ 处理出错: {str(e)}")
                import traceback
                traceback_str = traceback.format_exc()
                self.log(f"详细错误信息:\n{traceback_str}")

        # 处理完成
        self.update_progress(100)
        self.update_status("处理完成")

        # 统计结果
        result_message = f"处理完成！成功: {success_count} 个，失败: {failed_count} 个"
        self.log(f"{'=' * 60}")
        self.log(result_message)
        self.log(f"输出目录: {self.output_dir_var.get()}")

        # 显示完成消息
        messagebox.showinfo(
            "完成",
            f"{result_message}\n\n"
            f"输出目录: {self.output_dir_var.get()}\n\n"
            f"【一一对应结果】:\n"
            + "\n".join([f"{os.path.basename(self.target_files[i])} -> "
                         f"{os.path.splitext(os.path.basename(self.target_files[i]))[0]}-安全检测报告.docx"
                         for i in range(min(total_pairs, 10))])
            + ("\n..." if total_pairs > 10 else "")
        )

    def open_file(self, file_path):
        """打开文件"""
        try:
            os.startfile(file_path)  # Windows
            self.log(f"已打开文件: {os.path.basename(file_path)}")
        except:
            try:
                subprocess.call(['open', file_path])  # macOS
                self.log(f"已打开文件: {os.path.basename(file_path)}")
            except:
                try:
                    subprocess.call(['xdg-open', file_path])  # Linux
                    self.log(f"已打开文件: {os.path.basename(file_path)}")
                except:
                    self.log(f"无法自动打开文件: {os.path.basename(file_path)}")

    def generate_word_report(self, target_file, scan_file, output_file):
        """生成Word报告"""
        try:
            # 导入Word报告核心生成器
            from word_report_core import WordReportCore

            # 解析目标列表
            target_list = self.parse_target_list(target_file)
            if not target_list:
                self.log(f"错误: 目标列表文件 '{os.path.basename(target_file)}' 为空或无法解析")
                return False

            # 解析扫描结果
            scan_results = self.parse_nuclei_results(scan_file)
            if len(scan_results) == 0:
                self.log(f"警告: 扫描结果文件 '{os.path.basename(scan_file)}' 未找到任何扫描记录")

            # 创建IP报告数据
            report_core = WordReportCore(log_callback=self.log)
            report_data = report_core.create_ip_report_data(
                target_file, scan_file,
                self.scan_date_var.get(),
                self.tool_version_var.get(),
                []
            )

            if not report_data:
                self.log("创建报告数据失败")
                return False

            # 生成Word报告
            success = report_core.generate_word_report(
                "",  # 空模板，使用默认格式
                report_data,
                output_file
            )

            return success

        except Exception as e:
            self.log(f"生成Word报告时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False

    def generate_excel_report(self, target_file, scan_file, output_file):
        """生成Excel报告"""
        try:
            # 解析目标列表
            target_list = self.parse_target_list(target_file)
            if not target_list:
                self.log(f"错误: 目标列表文件 '{os.path.basename(target_file)}' 为空或无法解析")
                return False

            # 解析扫描结果
            scan_results = self.parse_nuclei_results(scan_file)
            if len(scan_results) == 0:
                self.log(f"警告: 扫描结果文件 '{os.path.basename(scan_file)}' 未找到任何扫描记录")

            # 创建汇总表格
            summary_data, target_display_names = self.create_summary_table(target_list, scan_results)

            # 创建合并后的详情表格
            vulnerability_data = self.create_merged_vulnerability_table(scan_results, target_display_names)

            # 生成Excel报告
            self.create_excel_report(summary_data, vulnerability_data, output_file)

            return True

        except Exception as e:
            self.log(f"生成Excel报告时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False

    def clean_target(self, target):
        """清理目标字符串，去除多余空格和特殊字符"""
        if not target:
            return ""

        # 移除开头和结尾的空白字符
        target = target.strip()

        # 移除末尾的特殊字符
        target = re.sub(r'[\[\]\{\}<>]$', '', target)

        # 修复IP地址中的空格：将 "172. 17. 0. 254" 转换为 "172.17.0.254"
        ip_pattern = r'(\d{1,3})\.\s*(\d{1,3})\.\s*(\d{1,3})\.\s*(\d{1,3})'

        def fix_ip_spaces(match):
            return f"{match.group(1)}.{match.group(2)}.{match.group(3)}.{match.group(4)}"

        target = re.sub(ip_pattern, fix_ip_spaces, target)

        # 移除URL中的多余空格
        if '://' in target:
            # 将 "https://192. 168. 0. 234" 转换为 "https://192.168.0.234"
            protocol, rest = target.split('://', 1)
            # 清理主机部分中的空格
            host = rest.split('/')[0] if '/' in rest else rest
            host = re.sub(r'\s+', '', host)  # 移除所有空格
            if '/' in rest:
                path = '/' + '/'.join(rest.split('/')[1:])
                target = f"{protocol}://{host}{path}"
            else:
                target = f"{protocol}://{host}"

        return target

    def parse_target_list(self, file_path):
        """解析目标列表文件"""
        targets = []
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):
                        # 清理目标
                        line = line.split('#')[0].strip()  # 移除行内注释
                        line = self.clean_target(line)
                        if line:  # 确保清理后不为空
                            targets.append(line)
        except Exception as e:
            self.log(f"解析目标列表时出错: {e}")
        return targets

    def parse_nuclei_results(self, file_path):
        """解析Nuclei扫描结果文件"""
        results = []

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        except:
            try:
                with open(file_path, 'r', encoding='gbk', errors='ignore') as f:
                    content = f.read()
            except Exception as e:
                self.log(f"读取文件时出错: {e}")
                return results

        # 逐行解析
        lines = content.split('\n')
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue

            # 跳过非标准行（包含乱码）
            if not line.startswith('[') or '] [' not in line:
                continue

            try:
                # 使用正则表达式提取各部分
                # 格式: [template] [protocol] [severity] target ["extra info"]
                pattern = r'\[([^\]]+)\] \[([^\]]+)\] \[([^\]]+)\] (.+?)(?: \["(.*)"\])?$'
                match = re.match(pattern, line)

                if match:
                    template, protocol, severity, target, extra_info = match.groups()

                    # 清理目标
                    target = self.clean_target(target)

                    results.append({
                        'template': template.strip(),
                        'protocol': protocol.strip(),
                        'severity': severity.strip(),
                        'target': target,
                        'extra_info': extra_info if extra_info else "",
                        'line_num': line_num
                    })
                else:
                    # 尝试更宽松的匹配
                    parts = line.split('] [', 3)
                    if len(parts) >= 3:
                        template = parts[0].strip('[')
                        protocol = parts[1]

                        # 提取严重程度
                        severity_part = parts[2]
                        if ']' in severity_part:
                            severity = severity_part.split(']')[0]
                            remaining = severity_part.split(']', 1)[1]
                            if len(parts) > 3:
                                remaining += ']' + ']'.join(parts[3:])
                        else:
                            severity = severity_part
                            remaining = ']'.join(parts[3:]) if len(parts) > 3 else ''

                        # 提取目标
                        target = remaining.strip()

                        # 清理目标
                        target = self.clean_target(target)

                        results.append({
                            'template': template.strip(),
                            'protocol': protocol.strip(),
                            'severity': severity.strip(),
                            'target': target,
                            'extra_info': "",
                            'line_num': line_num
                        })
            except Exception as e:
                self.log(f"警告: 解析第{line_num}行时出错: {e}")
                continue

        return results

    def normalize_target_display(self, target):
        """
        标准化目标地址用于显示
        - URL保持不变
        - IP:端口 只保留IP
        """
        if not target:
            return ""

        target = self.clean_target(target)

        # 检查是否是IP地址（可能带端口）
        ip_pattern = r'^(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})(:\d+)?$'
        match = re.match(ip_pattern, target)

        if match:
            # 纯IP地址，返回IP部分
            return match.group(1)

        # 检查是否是URL格式
        url_pattern = r'^(https?://)([^/]+)(/.*)?$'
        url_match = re.match(url_pattern, target)

        if url_match:
            protocol = url_match.group(1)
            host = url_match.group(2)
            path = url_match.group(3) or ""

            # 如果host是IP:端口，只保留IP
            host_match = re.match(ip_pattern, host)
            if host_match:
                host = host_match.group(1)

            return f"{protocol}{host}{path}"

        # 其他情况返回原样
        return target

    def get_target_key(self, target):
        """
        获取目标的匹配键
        用于在匹配时忽略协议和端口
        """
        if not target:
            return ""

        target = self.clean_target(target)

        # 移除协议
        if target.startswith('http://'):
            target = target[7:]
        elif target.startswith('https://'):
            target = target[8:]

        # 移除路径
        if '/' in target:
            target = target.split('/')[0]

        # 对于IP地址，移除端口
        ip_pattern = r'^(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})(:\d+)?$'
        match = re.match(ip_pattern, target)
        if match:
            return match.group(1)  # 只返回IP部分

        # 对于非IP，移除端口
        if ':' in target:
            target = target.split(':')[0]

        return target

    def categorize_severity(self, severity):
        """将严重程度分类"""
        severity_lower = severity.lower()

        if 'critical' in severity_lower or 'high' in severity_lower:
            return '高'
        elif 'medium' in severity_lower:
            return '中'
        elif 'low' in severity_lower:
            return '低'
        else:
            return '信息'

    def create_summary_table(self, target_list, scan_results):
        """创建汇总表格"""
        # 初始化所有目标
        target_stats = {}
        target_display_names = {}

        # 首先处理所有原始目标
        for original_target in target_list:
            display_name = self.normalize_target_display(original_target)
            target_key = self.get_target_key(original_target)

            if target_key:  # 确保key不为空
                target_display_names[target_key] = display_name
                target_stats[target_key] = {
                    'display_name': display_name,
                    '高': 0, '中': 0, '低': 0, '信息': 0, '小计': 0
                }

        # 处理扫描结果
        for result in scan_results:
            scan_target = result['target']
            target_key = self.get_target_key(scan_target)
            severity = self.categorize_severity(result['severity'])

            if not target_key:  # 如果key为空，跳过
                continue

            # 如果这个目标在目标列表中，增加计数
            if target_key in target_stats:
                target_stats[target_key][severity] += 1
                target_stats[target_key]['小计'] += 1
            else:
                # 如果不在目标列表中，添加到统计中
                display_name = self.normalize_target_display(scan_target)
                target_stats[target_key] = {
                    'display_name': display_name,
                    '高': 0, '中': 0, '低': 0, '信息': 0, '小计': 0
                }
                target_display_names[target_key] = display_name

                target_stats[target_key][severity] += 1
                target_stats[target_key]['小计'] += 1

        # 转换为表格格式
        summary_data = []
        for i, (target_key, stats) in enumerate(target_stats.items(), 1):
            summary_data.append({
                '序号': i,
                '检测目标': stats['display_name'],
                '高': stats['高'],
                '中': stats['中'],
                '低': stats['低'],
                '信息': stats['信息'],
                '小计': stats['小计']
            })

        # 排序：有漏洞的在前，按小计降序
        summary_data.sort(key=lambda x: (-x['小计'], x['检测目标']))

        # 重新编号
        for i, row in enumerate(summary_data, 1):
            row['序号'] = i

        return summary_data, target_display_names

    def create_merged_vulnerability_table(self, scan_results, target_display_names):
        """
        创建合并后的漏洞详情表格
        相同漏洞名称的目标合并到同一个单元格
        """
        # 使用字典来合并相同漏洞名称的记录
        vuln_dict = defaultdict(lambda: {
            'template': '',
            'severity': '',
            'targets': set(),  # 使用集合去重
        })

        for result in scan_results:
            scan_target = result['target']
            target_key = self.get_target_key(scan_target)
            template = result['template']
            severity = self.categorize_severity(result['severity'])

            if not target_key:  # 如果key为空，跳过
                continue

            # 获取显示名称
            if target_key in target_display_names:
                display_name = target_display_names[target_key]
            else:
                display_name = self.normalize_target_display(scan_target)

            # 添加到字典中
            if template not in vuln_dict:
                vuln_dict[template] = {
                    'template': template,
                    'severity': severity,
                    'targets': set(),
                }

            vuln_dict[template]['targets'].add(display_name)

        # 转换为列表格式
        vulnerability_data = []

        for i, (template, data) in enumerate(vuln_dict.items(), 1):
            # 将目标集合转换为排序后的列表，然后用逗号连接
            sorted_targets = sorted(data['targets'])
            targets_str = ', '.join(sorted_targets)

            vulnerability_data.append({
                '序号': i,
                '安全漏洞名称': template,
                '关联目标': targets_str,
                '严重程度': data['severity']
            })

        # 排序：按严重程度（高->中->低->信息），然后按漏洞名称
        severity_order = {'高': 0, '中': 1, '低': 2, '信息': 3}
        vulnerability_data.sort(key=lambda x: (severity_order.get(x['严重程度'], 4), x['安全漏洞名称']))

        # 重新编号
        for i, row in enumerate(vulnerability_data, 1):
            row['序号'] = i

        return vulnerability_data

    def create_excel_report(self, summary_data, vulnerability_data, output_file):
        """创建Excel报告"""
        # 创建DataFrame
        summary_df = pd.DataFrame(summary_data, columns=['序号', '检测目标', '高', '中', '低', '信息', '小计'])

        # 使用合并后的漏洞表格（只有4列）
        vuln_columns = ['序号', '安全漏洞名称', '关联目标', '严重程度']
        vulnerability_df = pd.DataFrame(vulnerability_data, columns=vuln_columns)

        # 创建Excel写入器
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 写入汇总表格
            summary_df.to_excel(writer, sheet_name='漏洞汇总', index=False)

            # 写入详细表格
            vulnerability_df.to_excel(writer, sheet_name='漏洞详情', index=False)

            # 获取工作簿和工作表
            workbook = writer.book
            summary_sheet = workbook['漏洞汇总']
            detail_sheet = workbook['漏洞详情']

            # 设置样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(vertical="top", wrap_text=True)  # 自动换行

            # 格式化汇总表头
            for col in range(1, 8):
                cell = summary_sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # 设置列宽
            summary_sheet.column_dimensions['A'].width = 10
            summary_sheet.column_dimensions['B'].width = 30
            for col in ['C', 'D', 'E', 'F', 'G']:
                summary_sheet.column_dimensions[col].width = 10

            # 格式化详情表头
            for col in range(1, 5):  # 现在只有4列
                cell = detail_sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # 设置列宽
            detail_sheet.column_dimensions['A'].width = 10  # 序号
            detail_sheet.column_dimensions['B'].width = 40  # 安全漏洞名称
            detail_sheet.column_dimensions['C'].width = 60  # 关联目标（合并后可能较长）
            detail_sheet.column_dimensions['D'].width = 15  # 严重程度

            # 为严重程度添加颜色
            severity_colors = {
                '高': 'FFC7CE',
                '中': 'FFEB9C',
                '低': 'C6EFCE',
                '信息': 'BDD7EE'
            }

            # 严重程度在第4列（D列）
            severity_col_index = 4

            for row in range(2, len(vulnerability_data) + 2):
                severity_cell = detail_sheet.cell(row=row, column=severity_col_index)
                severity = severity_cell.value
                if severity in severity_colors:
                    severity_cell.fill = PatternFill(
                        start_color=severity_colors[severity],
                        end_color=severity_colors[severity],
                        fill_type="solid"
                    )

                # 为关联目标列设置自动换行
                target_cell = detail_sheet.cell(row=row, column=3)  # C列
                target_cell.alignment = wrap_alignment

            # 为小计为0的行添加灰色背景
            for row in range(2, len(summary_data) + 2):
                total_cell = summary_sheet.cell(row=row, column=7)
                if total_cell.value == 0:
                    for col in range(1, 8):
                        cell = summary_sheet.cell(row=row, column=col)
                        cell.fill = PatternFill(
                            start_color="F2F2F2",
                            end_color="F2F2F2",
                            fill_type="solid"
                        )

            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 为汇总表添加边框
            for row in summary_sheet.iter_rows(min_row=1, max_row=len(summary_data) + 1, min_col=1, max_col=7):
                for cell in row:
                    cell.border = thin_border

            # 为详情表添加边框
            for row in detail_sheet.iter_rows(min_row=1, max_row=len(vulnerability_data) + 1, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
