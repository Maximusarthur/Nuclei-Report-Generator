#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
设备名漏洞报告生成器（完整版）
包含GUI界面和核心处理逻辑
"""

import re
import pandas as pd
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
from datetime import datetime
import subprocess


class DeviceReportGenerator:
    """设备名漏洞报告生成器（完整版）"""

    def __init__(self, root, main_app):
        self.root = root
        self.main_app = main_app
        self.root.title("设备名漏洞报告生成器")
        self.root.geometry("900x700")

        # 设置窗口图标
        try:
            self.root.iconbitmap(default="nuclei.ico")
        except:
            pass

        # 初始化变量
        self.device_info_files = []
        self.scan_report_files = []
        self.device_info_cache = {}  # 缓存设备信息，避免重复解析

        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        # 返回首页按钮
        back_frame = tk.Frame(self.root)
        back_frame.pack(fill="x", padx=10, pady=5)

        tk.Button(back_frame, text="← 返回首页",
                  font=("微软雅黑", 10), bg="#95a5a6", fg="white",
                  command=self.return_to_main).pack(anchor="w")

        # 标题
        title_label = tk.Label(self.root, text="设备名漏洞报告生成器",
                               font=("微软雅黑", 18, "bold"), fg="#2c3e50")
        title_label.pack(pady=10)

        # 说明标签
        desc_text = """• 设备信息文件和扫描报告文件按顺序一一对应
• 自动将IP地址转换为设备名称
• 支持批量处理多对文件
• 生成标准化的设备漏洞统计表"""
        desc_label = tk.Label(self.root, text=desc_text, font=("微软雅黑", 10),
                              justify="left", fg="#34495e", bg="#ecf0f1", wraplength=850)
        desc_label.pack(pady=5, padx=20, fill="x")

        # 创建主框架 - 使用PanedWindow实现可调整分割
        main_paned = tk.PanedWindow(self.root, orient=tk.VERTICAL, sashrelief=tk.RAISED)
        main_paned.pack(fill="both", expand=True, padx=10, pady=5)

        # 上半部分：文件选择和输出设置
        top_frame = tk.Frame(main_paned)
        main_paned.add(top_frame, minsize=400)

        # 使用Notebook标签页优化空间利用
        style = ttk.Style()
        style.configure('Custom.TNotebook.Tab',
                        font=('微软雅黑', 12, 'bold'),
                        padding=[15, 8],
                        background='#ecf0f1')

        style.map('Custom.TNotebook.Tab',
                  background=[('selected', '#9b59b6')],
                  foreground=[('selected', 'white')])

        notebook = ttk.Notebook(top_frame, style='Custom.TNotebook')
        notebook.pack(fill="both", expand=True, padx=5, pady=5)

        # 标签页1：文件选择 - 使用彩色边框
        self.file_content = tk.Frame(notebook, bg='white', highlightbackground='#9b59b6', highlightthickness=2)
        notebook.add(self.file_content, text="📁 文件选择")

        # 标签页2：输出和控制 - 使用彩色边框
        self.control_content = tk.Frame(notebook, bg='white', highlightbackground='#2ecc71', highlightthickness=2)
        notebook.add(self.control_content, text="⚙️ 输出控制")

        # 创建文件选择内容
        self.create_file_content()

        # 创建输出控制内容
        self.create_control_content()

        # 下半部分：进度条和日志
        bottom_frame = tk.Frame(main_paned)
        main_paned.add(bottom_frame, minsize=200)

        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(bottom_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", padx=10, pady=(10, 5))

        # 日志文本框
        log_frame = tk.LabelFrame(bottom_frame, text="处理日志", font=("微软雅黑", 11, "bold"))
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.log_text = scrolledtext.ScrolledText(log_frame, height=6, font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, padx=10, pady=10)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = tk.Label(self.root, textvariable=self.status_var,
                              font=("微软雅黑", 9), bg="#34495e", fg="white",
                              anchor="w", padx=10)
        status_bar.pack(side="bottom", fill="x")

        # 设置窗口最小尺寸
        self.root.minsize(800, 600)

        # 添加日志
        self.log("设备名漏洞报告生成器启动成功")
        self.log(f"当前工作目录: {os.getcwd()}")

    def create_file_content(self):
        """创建文件选择内容"""
        # 左侧设备文件，右侧扫描文件
        file_select_frame = tk.Frame(self.file_content)
        file_select_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 设备信息文件选择
        device_frame = tk.LabelFrame(file_select_frame, text="设备信息文件(.txt)",
                                     font=("微软雅黑", 10, "bold"))
        device_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # 设备文件列表框
        self.device_listbox = tk.Listbox(device_frame, height=8, font=("Consolas", 9))
        self.device_listbox.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        device_scroll = tk.Scrollbar(device_frame)
        device_scroll.pack(side="right", fill="y")
        self.device_listbox.config(yscrollcommand=device_scroll.set)
        device_scroll.config(command=self.device_listbox.yview)

        # 设备文件按钮
        device_btn_frame = tk.Frame(device_frame)
        device_btn_frame.pack(fill="x", padx=5, pady=5)

        tk.Button(device_btn_frame, text="添加", command=self.add_device_file,
                  font=("微软雅黑", 9), bg="#9b59b6", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(device_btn_frame, text="移除", command=self.remove_selected_device,
                  font=("微软雅黑", 9), bg="#e74c3c", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(device_btn_frame, text="清空", command=self.clear_device_list,
                  font=("微软雅黑", 9), bg="#95a5a6", fg="white", width=8).pack(side="left", padx=2)

        # 扫描报告文件选择
        scan_frame = tk.LabelFrame(file_select_frame, text="扫描报告文件(.txt)",
                                   font=("微软雅黑", 10, "bold"))
        scan_frame.pack(side="right", fill="both", expand=True, padx=(5, 0))

        # 扫描文件列表框
        self.scan_listbox = tk.Listbox(scan_frame, height=8, font=("Consolas", 9))
        self.scan_listbox.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        scan_scroll = tk.Scrollbar(scan_frame)
        scan_scroll.pack(side="right", fill="y")
        self.scan_listbox.config(yscrollcommand=scan_scroll.set)
        scan_scroll.config(command=self.scan_listbox.yview)

        # 扫描文件按钮
        scan_btn_frame = tk.Frame(scan_frame)
        scan_btn_frame.pack(fill="x", padx=5, pady=5)

        tk.Button(scan_btn_frame, text="添加", command=self.add_scan_file,
                  font=("微软雅黑", 9), bg="#3498db", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(scan_btn_frame, text="移除", command=self.remove_selected_scan,
                  font=("微软雅黑", 9), bg="#e74c3c", fg="white", width=8).pack(side="left", padx=2)
        tk.Button(scan_btn_frame, text="清空", command=self.clear_scan_list,
                  font=("微软雅黑", 9), bg="#95a5a6", fg="white", width=8).pack(side="left", padx=2)

        # 文件对应提示
        self.pair_label = tk.Label(self.file_content, text="📝 请添加文件",
                                   font=("微软雅黑", 9), fg="#95a5a6")
        self.pair_label.pack(anchor="w", pady=(10, 5), padx=15)

        # 绑定列表变化事件
        self.device_listbox.bind('<<ListboxSelect>>', self.update_pair_status)
        self.scan_listbox.bind('<<ListboxSelect>>', self.update_pair_status)

    def create_control_content(self):
        """创建输出控制内容"""
        # 输出目录选择
        tk.Label(self.control_content, text="输出目录:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(10, 5), padx=15)

        output_frame = tk.Frame(self.control_content)
        output_frame.pack(fill="x", padx=15, pady=5)

        self.output_dir_var = tk.StringVar(value=os.path.join(os.getcwd(), "device_reports"))
        output_entry = tk.Entry(output_frame, textvariable=self.output_dir_var,
                                font=("微软雅黑", 9))
        output_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))

        tk.Button(output_frame, text="浏览", command=self.select_output_dir,
                  font=("微软雅黑", 9), bg="#2ecc71", fg="white", width=8).pack(side="right")

        # 检测时间输入
        tk.Label(self.control_content, text="检测时间:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(20, 5), padx=15)

        # 使用当前日期作为默认值
        default_date = datetime.now().strftime("%Y年%m月%d日")
        self.scan_date_var = tk.StringVar(value=default_date)

        date_frame = tk.Frame(self.control_content)
        date_frame.pack(fill="x", padx=15, pady=5)

        date_entry = tk.Entry(date_frame, textvariable=self.scan_date_var,
                              font=("微软雅黑", 9))
        date_entry.pack(side="left", fill="x", expand=True)

        tk.Label(date_frame, text="格式：YYYY年MM月DD日",
                 font=("微软雅黑", 8), fg="#7f8c8d").pack(side="right", padx=(5, 0))

        # 版本号输入
        tk.Label(self.control_content, text="Nuclei版本号:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(15, 5), padx=15)

        self.tool_version_var = tk.StringVar(value="V3.5.1,nuclei-templates v10.3.2")

        version_frame = tk.Frame(self.control_content)
        version_frame.pack(fill="x", padx=15, pady=5)

        version_entry = tk.Entry(version_frame, textvariable=self.tool_version_var,
                                 font=("微软雅黑", 9))
        version_entry.pack(side="left", fill="x", expand=True)

        tk.Label(version_frame, text="示例：V3.5.1,nuclei-templates v10.3.2",
                 font=("微软雅黑", 8), fg="#7f8c8d").pack(side="right", padx=(5, 0))

        # 报告类型选择
        tk.Label(self.control_content, text="报告格式:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(20, 5), padx=15)

        self.report_type_var = tk.StringVar(value="excel")

        report_frame = tk.Frame(self.control_content)
        report_frame.pack(fill="x", padx=15, pady=5)

        tk.Radiobutton(report_frame, text="Word文档(.docx)",
                       variable=self.report_type_var, value="word",
                       font=("微软雅黑", 9)).pack(side="left", padx=(0, 15))

        tk.Radiobutton(report_frame, text="Excel表格(.xlsx)",
                       variable=self.report_type_var, value="excel",
                       font=("微软雅黑", 9)).pack(side="left")

        # 处理选项
        tk.Label(self.control_content, text="处理选项:",
                 font=("微软雅黑", 10)).pack(anchor="w", pady=(20, 5), padx=15)

        self.auto_open_var = tk.BooleanVar(value=True)
        tk.Checkbutton(self.control_content, text="处理后自动打开报告文件",
                       variable=self.auto_open_var, font=("微软雅黑", 9)).pack(anchor="w", padx=15)

        # 处理按钮
        btn_frame = tk.Frame(self.control_content)
        btn_frame.pack(fill="x", pady=30, padx=15)

        self.process_btn = tk.Button(btn_frame, text="开始生成报告 (一一对应)",
                                     command=self.start_processing,
                                     font=("微软雅黑", 12, "bold"), bg="#27ae60", fg="white",
                                     padx=20, pady=10, state=tk.NORMAL)
        self.process_btn.pack(fill="x", pady=5)

    def return_to_main(self):
        """返回主界面"""
        self.root.destroy()
        self.main_app.return_to_home()

    def add_device_file(self):
        """添加设备信息文件"""
        files = filedialog.askopenfilenames(
            title="选择设备信息文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if files:
            for file in files:
                if file not in self.device_info_files:
                    self.device_info_files.append(file)
                    self.device_listbox.insert(tk.END, os.path.basename(file))
                    self.log(f"添加设备信息文件: {os.path.basename(file)}")
            self.update_pair_status()

    def remove_selected_device(self):
        """移除选中的设备信息文件"""
        selection = self.device_listbox.curselection()
        if selection:
            index = selection[0]
            removed_file = self.device_info_files.pop(index)
            self.device_listbox.delete(index)
            # 从缓存中移除
            if removed_file in self.device_info_cache:
                del self.device_info_cache[removed_file]
            self.log(f"移除设备信息文件: {os.path.basename(removed_file)}")
            self.update_pair_status()

    def clear_device_list(self):
        """清空设备信息文件列表"""
        self.device_info_files.clear()
        self.device_info_cache.clear()
        self.device_listbox.delete(0, tk.END)
        self.log("清空设备信息文件列表")
        self.update_pair_status()

    def add_scan_file(self):
        """添加扫描报告文件"""
        files = filedialog.askopenfilenames(
            title="选择扫描报告文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if files:
            for file in files:
                if file not in self.scan_report_files:
                    self.scan_report_files.append(file)
                    self.scan_listbox.insert(tk.END, os.path.basename(file))
                    self.log(f"添加扫描报告文件: {os.path.basename(file)}")
            self.update_pair_status()

    def remove_selected_scan(self):
        """移除选中的扫描报告文件"""
        selection = self.scan_listbox.curselection()
        if selection:
            index = selection[0]
            removed_file = self.scan_report_files.pop(index)
            self.scan_listbox.delete(index)
            self.log(f"移除扫描报告文件: {os.path.basename(removed_file)}")
            self.update_pair_status()

    def clear_scan_list(self):
        """清空扫描报告文件列表"""
        self.scan_report_files.clear()
        self.scan_listbox.delete(0, tk.END)
        self.log("清空扫描报告文件列表")
        self.update_pair_status()

    def update_pair_status(self, event=None):
        """更新文件对应关系状态"""
        device_count = len(self.device_info_files)
        scan_count = len(self.scan_report_files)

        if device_count == scan_count:
            if device_count == 0:
                self.pair_label.config(text="📝 请添加文件", fg="#95a5a6")
                self.process_btn.config(state=tk.NORMAL)
            else:
                self.pair_label.config(text=f"✅ 文件对应关系正常 ({device_count} 对文件)", fg="#27ae60")
                self.process_btn.config(state=tk.NORMAL)
        else:
            self.pair_label.config(
                text=f"⚠️ 文件数量不匹配: 设备文件 {device_count} 个，扫描文件 {scan_count} 个",
                fg="#e74c3c"
            )
            self.process_btn.config(state=tk.DISABLED)

    def select_output_dir(self):
        """选择输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir_var.set(directory)
            self.log(f"设置输出目录: {directory}")

    def log(self, message):
        """添加日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_status(self, message):
        """更新状态栏"""
        self.status_var.set(message)
        self.root.update_idletasks()

    def update_progress(self, value):
        """更新进度条"""
        self.progress_var.set(value)
        self.root.update_idletasks()

    def start_processing(self):
        """开始批量处理"""
        # 检查文件数量是否匹配
        if len(self.device_info_files) != len(self.scan_report_files):
            messagebox.showerror(
                "错误",
                f"文件数量不匹配！\n\n"
                f"设备信息文件数: {len(self.device_info_files)}\n"
                f"扫描报告文件数: {len(self.scan_report_files)}\n\n"
                f"请确保两个列表的数量相同且顺序对应。"
            )
            return

        if not self.device_info_files or not self.scan_report_files:
            messagebox.showwarning("警告", "请至少添加一对文件")
            return

        # 检查输出目录
        output_dir = self.output_dir_var.get()
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
            except Exception as e:
                messagebox.showerror("错误", f"无法创建输出目录: {e}")
                return

        # 在后台线程中处理
        processing_thread = threading.Thread(target=self.process_files_one_to_one)
        processing_thread.daemon = True
        processing_thread.start()

    def process_files_one_to_one(self):
        """一一对应处理文件"""
        total_pairs = len(self.device_info_files)

        self.update_status("开始处理文件...")
        self.update_progress(0)
        self.log(f"开始一一对应处理 {total_pairs} 对文件")

        success_count = 0
        failed_count = 0

        for i in range(total_pairs):
            try:
                device_file = self.device_info_files[i]
                scan_file = self.scan_report_files[i]

                # 计算进度
                progress = ((i + 1) / total_pairs) * 100
                self.update_progress(progress)

                # 获取文件名
                device_basename = os.path.splitext(os.path.basename(device_file))[0]
                scan_basename = os.path.splitext(os.path.basename(scan_file))[0]

                # 生成输出文件名
                report_type = self.report_type_var.get()
                if report_type == "word":
                    output_filename = f"{device_basename}-安全检测报告.docx"
                else:
                    output_filename = f"{device_basename}-设备报告.xlsx"

                output_path = os.path.join(self.output_dir_var.get(), output_filename)

                # 记录处理开始
                pair_info = f"配对 {i + 1}/{total_pairs}: {os.path.basename(device_file)} -> {os.path.basename(scan_file)}"
                self.log(f"开始处理: {pair_info}")
                self.update_status(f"处理中: {output_filename}")

                # 调用处理函数
                if report_type == "word":
                    success = self.generate_word_report(device_file, scan_file, output_path)
                else:
                    success = self.generate_excel_report(device_file, scan_file, output_path)

                if success:
                    success_count += 1
                    self.log(f"✓ 处理完成: {output_filename}")

                    # 自动打开文件
                    if self.auto_open_var.get():
                        self.open_file(output_path)
                else:
                    failed_count += 1
                    self.log(f"✗ 处理失败: {os.path.basename(scan_file)}")

            except Exception as e:
                failed_count += 1
                self.log(f"✗ 处理出错: {str(e)}")
                import traceback
                traceback_str = traceback.format_exc()
                self.log(f"详细错误信息:\n{traceback_str}")

        # 处理完成
        self.update_progress(100)
        self.update_status("处理完成")

        # 统计结果
        result_message = f"处理完成！成功: {success_count} 个，失败: {failed_count} 个"
        self.log(f"{'=' * 60}")
        self.log(result_message)
        self.log(f"输出目录: {self.output_dir_var.get()}")

        # 显示完成消息
        messagebox.showinfo(
            "完成",
            f"{result_message}\n\n"
            f"输出目录: {self.output_dir_var.get()}\n\n"
            f"【一一对应结果】:\n"
            + "\n".join([f"{os.path.basename(self.device_info_files[i])} -> "
                         f"{os.path.splitext(os.path.basename(self.device_info_files[i]))[0]}-安全检测报告.docx"
                         for i in range(min(total_pairs, 10))])
            + ("\n..." if total_pairs > 10 else "")
        )

    def open_file(self, file_path):
        """打开文件"""
        try:
            os.startfile(file_path)  # Windows
            self.log(f"已打开文件: {os.path.basename(file_path)}")
        except:
            try:
                subprocess.call(['open', file_path])  # macOS
                self.log(f"已打开文件: {os.path.basename(file_path)}")
            except:
                try:
                    subprocess.call(['xdg-open', file_path])  # Linux
                    self.log(f"已打开文件: {os.path.basename(file_path)}")
                except:
                    self.log(f"无法自动打开文件: {os.path.basename(file_path)}")

    def generate_word_report(self, device_file, scan_file, output_file):
        """生成Word报告"""
        try:
            # 导入Word报告核心生成器
            from word_report_core import WordReportCore

            # 解析设备信息
            device_info = self.parse_device_info(device_file)
            if not device_info:
                self.log(f"错误: 设备信息文件 '{os.path.basename(device_file)}' 为空或无法解析")
                return False

            # 解析扫描结果
            scan_results = self.parse_nuclei_results(scan_file)
            if len(scan_results) == 0:
                self.log(f"警告: 扫描报告文件 '{os.path.basename(scan_file)}' 未找到任何扫描记录")

            # 创建设备报告数据
            report_core = WordReportCore(log_callback=self.log)
            report_data = report_core.create_device_report_data(
                device_file, scan_file,
                self.scan_date_var.get(),
                self.tool_version_var.get(),
                []
            )

            if not report_data:
                self.log("创建报告数据失败")
                return False

            # 生成Word报告
            success = report_core.generate_word_report(
                "",  # 空模板，使用默认格式
                report_data,
                output_file
            )

            return success

        except Exception as e:
            self.log(f"生成Word报告时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False

    def generate_excel_report(self, device_file, scan_file, output_file):
        """生成Excel报告"""
        try:
            # 解析设备信息
            device_info = self.parse_device_info(device_file)
            if not device_info:
                self.log(f"错误: 设备信息文件 '{os.path.basename(device_file)}' 为空或无法解析")
                return False

            # 解析扫描结果
            scan_results = self.parse_nuclei_results(scan_file)
            if len(scan_results) == 0:
                self.log(f"警告: 扫描报告文件 '{os.path.basename(scan_file)}' 未找到任何扫描记录")

            # 创建设备漏洞统计
            device_stats_data = self.create_device_statistics(device_info, scan_results)

            # 创建漏洞详情
            vulnerability_details_data = self.create_vulnerability_details(device_info, scan_results)

            # 生成Excel报告
            self.create_device_excel_report(device_stats_data, vulnerability_details_data, output_file)

            return True

        except Exception as e:
            self.log(f"生成Excel报告时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False

    def parse_device_info(self, file_path):
        """解析设备信息文件"""
        # 检查缓存
        if file_path in self.device_info_cache:
            return self.device_info_cache[file_path]

        device_info = {}
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        # 使用制表符分割
                        parts = line.split('\t')
                        if len(parts) >= 3:
                            device_name = parts[0].strip()
                            system_version = parts[1].strip()
                            ip = parts[2].strip()

                            # 清理IP地址
                            ip = self.clean_ip(ip)
                            if ip:
                                device_info[ip] = {
                                    'device_name': device_name,
                                    'system_version': system_version
                                }
            self.log(f"解析设备信息完成，共{len(device_info)}个设备")
            # 存入缓存
            self.device_info_cache[file_path] = device_info
        except Exception as e:
            self.log(f"解析设备信息文件时出错: {e}")
            raise e
        return device_info

    def clean_ip(self, ip):
        """清理IP地址"""
        if not ip:
            return ""

        ip = ip.strip()
        # 修复IP地址中的空格：将 "172. 17. 0. 254" 转换为 "172.17.0.254"
        ip_pattern = r'(\d{1,3})\.\s*(\d{1,3})\.\s*(\d{1,3})\.\s*(\d{1,3})'

        def fix_ip_spaces(match):
            return f"{match.group(1)}.{match.group(2)}.{match.group(3)}.{match.group(4)}"

        ip = re.sub(ip_pattern, fix_ip_spaces, ip)
        return ip

    def parse_nuclei_results(self, file_path):
        """解析Nuclei扫描结果文件"""
        results = []

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        except:
            try:
                with open(file_path, 'r', encoding='gbk', errors='ignore') as f:
                    content = f.read()
            except Exception as e:
                self.log(f"读取文件时出错: {e}")
                return results

        # 逐行解析
        lines = content.split('\n')
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue

            # 跳过非标准行（包含乱码）
            if not line.startswith('[') or '] [' not in line:
                continue

            try:
                # 使用正则表达式提取各部分
                # 格式: [template] [protocol] [severity] target ["extra info"]
                pattern = r'\[([^\]]+)\] \[([^\]]+)\] \[([^\]]+)\] (.+?)(?: \["(.*)"\])?$'
                match = re.match(pattern, line)

                if match:
                    template, protocol, severity, target, extra_info = match.groups()

                    # 清理目标
                    target = self.clean_target(target)

                    results.append({
                        'template': template.strip(),
                        'protocol': protocol.strip(),
                        'severity': severity.strip(),
                        'target': target,
                        'extra_info': extra_info if extra_info else "",
                        'line_num': line_num
                    })
                else:
                    # 尝试更宽松的匹配
                    parts = line.split('] [', 3)
                    if len(parts) >= 3:
                        template = parts[0].strip('[')
                        protocol = parts[1]

                        # 提取严重程度
                        severity_part = parts[2]
                        if ']' in severity_part:
                            severity = severity_part.split(']')[0]
                            remaining = severity_part.split(']', 1)[1]
                            if len(parts) > 3:
                                remaining += ']' + ']'.join(parts[3:])
                        else:
                            severity = severity_part
                            remaining = ']'.join(parts[3:]) if len(parts) > 3 else ''

                        # 提取目标
                        target = remaining.strip()

                        # 清理目标
                        target = self.clean_target(target)

                        results.append({
                            'template': template.strip(),
                            'protocol': protocol.strip(),
                            'severity': severity.strip(),
                            'target': target,
                            'extra_info': "",
                            'line_num': line_num
                        })
            except Exception as e:
                self.log(f"警告: 解析第{line_num}行时出错: {e}")
                continue

        return results

    def clean_target(self, target):
        """清理目标字符串"""
        if not target:
            return ""

        # 移除开头和结尾的空白字符
        target = target.strip()

        # 移除末尾的特殊字符
        target = re.sub(r'[\[\]\{\}<>]$', '', target)

        # 修复IP地址中的空格
        ip_pattern = r'(\d{1,3})\.\s*(\d{1,3})\.\s*(\d{1,3})\.\s*(\d{1,3})'

        def fix_ip_spaces(match):
            return f"{match.group(1)}.{match.group(2)}.{match.group(3)}.{match.group(4)}"

        target = re.sub(ip_pattern, fix_ip_spaces, target)

        # 移除URL中的多余空格
        if '://' in target:
            protocol, rest = target.split('://', 1)
            host = rest.split('/')[0] if '/' in rest else rest
            host = re.sub(r'\s+', '', host)
            if '/' in rest:
                path = '/' + '/'.join(rest.split('/')[1:])
                target = f"{protocol}://{host}{path}"
            else:
                target = f"{protocol}://{host}"

        return target

    def get_ip_from_target(self, target):
        """从目标中提取IP地址"""
        target = self.clean_target(target)

        # 如果是URL，提取主机部分
        if '://' in target:
            protocol, rest = target.split('://', 1)
            host = rest.split('/')[0] if '/' in rest else rest
        else:
            host = target

        # 提取IP部分（去除端口）
        if ':' in host:
            host = host.split(':')[0]

        # 检查是否是IP地址
        ip_pattern = r'^(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})$'
        match = re.match(ip_pattern, host)
        if match:
            return match.group(1)

        return None

    def categorize_severity(self, severity):
        """将严重程度分类"""
        severity_lower = severity.lower()

        if 'critical' in severity_lower or 'high' in severity_lower:
            return '高'
        elif 'medium' in severity_lower:
            return '中'
        elif 'low' in severity_lower:
            return '低'
        else:
            return '信息'

    def create_device_statistics(self, device_info, scan_results):
        """创建设备漏洞统计"""
        # 初始化设备统计字典
        device_stats = {}

        # 初始化所有设备
        for ip, info in device_info.items():
            device_stats[ip] = {
                'device_name': info['device_name'],
                'system_version': info['system_version'],
                '高': 0, '中': 0, '低': 0, '信息': 0, '小计': 0
            }

        # 处理扫描结果
        for result in scan_results:
            target = result['target']
            ip = self.get_ip_from_target(target)
            severity = self.categorize_severity(result['severity'])

            if ip and ip in device_stats:
                device_stats[ip][severity] += 1
                device_stats[ip]['小计'] += 1

        # 转换为表格格式
        stats_data = []
        for i, (ip, stats) in enumerate(device_stats.items(), 1):
            stats_data.append({
                '序号': i,
                '设备名称': stats['device_name'],
                '系统及版本': stats['system_version'],
                '高': stats['高'],
                '中': stats['中'],
                '低': stats['低'],
                '信息': stats['信息'],
                '小计': stats['小计']
            })

        # 排序：有漏洞的在前，按小计降序
        stats_data.sort(key=lambda x: (-x['小计'], x['设备名称']))

        # 重新编号
        for i, row in enumerate(stats_data, 1):
            row['序号'] = i

        return stats_data

    def create_vulnerability_details(self, device_info, scan_results):
        """创建漏洞详情"""
        # 使用字典来合并相同漏洞名称的记录
        vuln_dict = defaultdict(lambda: {
            'severity': '',
            'devices': set(),  # 使用集合去重
        })

        for result in scan_results:
            template = result['template']
            severity = self.categorize_severity(result['severity'])
            target = result['target']
            ip = self.get_ip_from_target(target)

            if ip and ip in device_info:
                device_name = device_info[ip]['device_name']
            else:
                # 如果找不到设备信息，使用IP
                device_name = ip if ip else target

            # 添加到字典中
            if template not in vuln_dict:
                vuln_dict[template] = {
                    'severity': severity,
                    'devices': set(),
                }

            vuln_dict[template]['devices'].add(device_name)

        # 转换为列表格式
        vulnerability_data = []

        for i, (template, data) in enumerate(vuln_dict.items(), 1):
            # 将设备集合转换为排序后的列表，然后用逗号连接
            sorted_devices = sorted(data['devices'])
            devices_str = ', '.join(sorted_devices)

            vulnerability_data.append({
                '序号': i,
                '安全漏洞名称': template,
                '关联目标': devices_str,
                '严重程度': data['severity']
            })

        # 排序：按严重程度（高->中->低->信息），然后按漏洞名称
        severity_order = {'高': 0, '中': 1, '低': 2, '信息': 3}
        vulnerability_data.sort(key=lambda x: (severity_order.get(x['严重程度'], 4), x['安全漏洞名称']))

        # 重新编号
        for i, row in enumerate(vulnerability_data, 1):
            row['序号'] = i

        return vulnerability_data

    def create_device_excel_report(self, device_stats, vulnerability_details, output_file):
        """创建设备Excel报告"""
        # 创建DataFrame
        stats_columns = ['序号', '设备名称', '系统及版本', '高', '中', '低', '信息', '小计']
        stats_df = pd.DataFrame(device_stats, columns=stats_columns)

        # 漏洞详情表格
        vuln_columns = ['序号', '安全漏洞名称', '关联目标', '严重程度']
        vulnerability_df = pd.DataFrame(vulnerability_details, columns=vuln_columns)

        # 创建Excel写入器
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 写入设备统计表格
            stats_df.to_excel(writer, sheet_name='设备漏洞统计', index=False)

            # 写入漏洞详情表格
            vulnerability_df.to_excel(writer, sheet_name='漏洞详情', index=False)

            # 获取工作簿和工作表
            workbook = writer.book
            stats_sheet = workbook['设备漏洞统计']
            detail_sheet = workbook['漏洞详情']

            # 设置样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(vertical="top", wrap_text=True)  # 自动换行

            # 格式化设备统计表头
            for col in range(1, 9):  # 8列
                cell = stats_sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # 设置列宽
            stats_sheet.column_dimensions['A'].width = 10  # 序号
            stats_sheet.column_dimensions['B'].width = 25  # 设备名称
            stats_sheet.column_dimensions['C'].width = 30  # 系统及版本
            for col in ['D', 'E', 'F', 'G', 'H']:
                stats_sheet.column_dimensions[col].width = 10

            # 格式化详情表头
            for col in range(1, 5):  # 4列
                cell = detail_sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # 设置列宽
            detail_sheet.column_dimensions['A'].width = 10  # 序号
            detail_sheet.column_dimensions['B'].width = 40  # 安全漏洞名称
            detail_sheet.column_dimensions['C'].width = 60  # 关联目标（合并后可能较长）
            detail_sheet.column_dimensions['D'].width = 15  # 严重程度

            # 为严重程度添加颜色
            severity_colors = {
                '高': 'FFC7CE',
                '中': 'FFEB9C',
                '低': 'C6EFCE',
                '信息': 'BDD7EE'
            }

            # 严重程度在第4列（D列）
            severity_col_index = 4

            for row in range(2, len(vulnerability_details) + 2):
                severity_cell = detail_sheet.cell(row=row, column=severity_col_index)
                severity = severity_cell.value
                if severity in severity_colors:
                    severity_cell.fill = PatternFill(
                        start_color=severity_colors[severity],
                        end_color=severity_colors[severity],
                        fill_type="solid"
                    )

                # 为关联目标列设置自动换行
                target_cell = detail_sheet.cell(row=row, column=3)  # C列
                target_cell.alignment = wrap_alignment

            # 为小计为0的行添加灰色背景
            for row in range(2, len(device_stats) + 2):
                total_cell = stats_sheet.cell(row=row, column=8)  # H列
                if total_cell.value == 0:
                    for col in range(1, 9):
                        cell = stats_sheet.cell(row=row, column=col)
                        cell.fill = PatternFill(
                            start_color="F2F2F2",
                            end_color="F2F2F2",
                            fill_type="solid"
                        )

            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 为设备统计表添加边框
            for row in stats_sheet.iter_rows(min_row=1, max_row=len(device_stats) + 1, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border

            # 为详情表添加边框
            for row in detail_sheet.iter_rows(min_row=1, max_row=len(vulnerability_details) + 1, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
